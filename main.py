import openpyxl
import win32com.client
import time
import os

dir = os.getcwd()

namaFile = input('Masukan Nama File Excel (Tanpa Format .xlsx): ')
rangeData = int(input('Berapa baris data yang diperlukan? (Silakan buka file sumber Anda) '))

weight1 = int(input('Masukan Nilai Bobot Lama Studi (Skala 1-5): '))
scale1 = input('Apakah menggunakan MAX atau MIN? (Ketik 1 untuk MAX dan 0 untuk MIN) ')

weight2 = int(input('Masukan Nilai Bobot IPK (Skala 1-5): '))
scale2 = input('Apakah menggunakan MAX atau MIN? (Ketik 1 untuk MAX dan 0 untuk MIN) ')

weight3 = int(input('Masukan Nilai Bobot Penghasilan Orang Tua (Skala 1-5): '))
scale3 = input('Apakah menggunakan MAX atau MIN? (Ketik 1 untuk MAX dan 0 untuk MIN) ')

weight4 = int(input('Masukan Nilai Bobot Jarak Rumah (Skala 1-5): '))
scale4 = input('Apakah menggunakan MAX atau MIN? (Ketik 1 untuk MAX dan 0 untuk MIN) ')

wb = openpyxl.load_workbook(f'{namaFile}.xlsx')
ws = wb.active
listSheet = wb.sheetnames
eq = "="
slash = "/"

# Fungsi Max dan Min
def formulaMAX(formula, letter_source):
    exe = str(formula)+slash+f"MAX({letter_source}2:{letter_source}{rangeData})"
    return exe

def formulaMIN(formula, letter_source):
    exe = f"MIN({letter_source}2:{letter_source}{rangeData})"+slash+str(formula)
    return exe

def measure(collumn_source, scale, letter_source):
    if scale == "1":
        exe = formulaMAX(collumn_source, letter_source)
        return exe
    elif scale == "0":
        exe = formulaMIN(collumn_source, letter_source)
        return exe
    else:
        print("Terjadi error!")

# Normalisasi
for i in range(rangeData):
    letter_source = "C"
    letter_target = "G"
    head_target = ws[f'{letter_target}1'] = "Lama Studi"
    collumn_source = ws[f'{letter_source}{i+2}'].value
    if scale1 == "1":
        calc = eq+formulaMAX(collumn_source, letter_source)
        if collumn_source is None:
            collumn_source = 1
            collumn_target = ws[f'{letter_target}{i+2}'].value = calc
        else:
            collumn_target = ws[f'{letter_target}{i+2}'].value = calc
    elif scale1 == "0":
        calc = eq+formulaMIN(collumn_source, letter_source)
        if collumn_source is None:
            collumn_source = 1
            collumn_target = ws[f'{letter_target}{i+2}'].value = calc
        else:
            collumn_target = ws[f'{letter_target}{i+2}'].value = calc
    else:
        print("Terjadi error!")
       
for i in range(rangeData):
    letter_source = "D"
    letter_target = "H"
    head_target = ws[f'{letter_target}1'] = "IPK"
    collumn_source = ws[f'{letter_source}{i+2}'].value
    if scale2 == "1":
        calc = eq+formulaMAX(collumn_source, letter_source)
        if collumn_source is None:
            collumn_source = 1
            collumn_target = ws[f'{letter_target}{i+2}'].value = calc
        else:
            collumn_target = ws[f'{letter_target}{i+2}'].value = calc
    elif scale2 == "0":
        calc = eq+formulaMIN(collumn_source, letter_source)
        if collumn_source is None:
            collumn_source = 1
            collumn_target = ws[f'{letter_target}{i+2}'].value = calc
        else:
            collumn_target = ws[f'{letter_target}{i+2}'].value = calc
    else:
        print("Terjadi error!")


for i in range(rangeData):
    letter_source = "E"
    letter_target = "I"
    head_target = ws[f'{letter_target}1'] = "Penghasilan Orang Tua"
    collumn_source = ws[f'{letter_source}{i+2}'].value
    if scale3 == "1":
        calc = eq+formulaMAX(collumn_source, letter_source)
        if collumn_source is None:
            collumn_source = 1
            collumn_target = ws[f'{letter_target}{i+2}'].value = calc
        else:
            collumn_target = ws[f'{letter_target}{i+2}'].value = calc
    elif scale3 == "0":
        calc = eq+formulaMIN(collumn_source, letter_source)
        if collumn_source is None:
            collumn_source = 1
            collumn_target = ws[f'{letter_target}{i+2}'].value = calc
        else:
            collumn_target = ws[f'{letter_target}{i+2}'].value = calc
    else:
        print("Terjadi error!")


for i in range(rangeData):
    letter_source = "F"
    letter_target = "J"
    head_target = ws[f'{letter_target}1'] = "Jarak Rumah"
    collumn_source = ws[f'{letter_source}{i+2}'].value
    if scale4 == "1":
        calc = eq+formulaMAX(collumn_source, letter_source)
        if collumn_source is None:
            collumn_source = 1
            collumn_target = ws[f'{letter_target}{i+2}'].value = calc
        else:
            collumn_target = ws[f'{letter_target}{i+2}'].value = calc
    elif scale4 == "0":
        calc = eq+formulaMIN(collumn_source, letter_source)
        if collumn_source is None:
            collumn_source = 1
            collumn_target = ws[f'{letter_target}{i+2}'].value = calc
        else:
            collumn_target = ws[f'{letter_target}{i+2}'].value = calc
    else:
        print("Terjadi error!")

# Preferensi
for i in range(rangeData):
    weightTotal = weight1+weight2+weight3+weight4
    letter_source1 = "C"
    letter_source2 = "D"
    letter_source3 = "E"
    letter_source4 = "F"
    letter_target = "K"
    head_target = ws[f'{letter_target}1'] = "Preferensi"
    collumn_source1 = ws[f'{letter_source1}{i+2}'].value
    collumn_source2 = ws[f'{letter_source2}{i+2}'].value
    collumn_source3 = ws[f'{letter_source3}{i+2}'].value
    collumn_source4 = ws[f'{letter_source4}{i+2}'].value
    measure1 = "(" + str(measure(collumn_source1, scale1, letter_source1)) + ")" + "*" + str(weight1/weightTotal) + "+"
    measure2 = "(" + str(measure(collumn_source2, scale2, letter_source2)) + ")" + "*" + str(weight2/weightTotal) + "+"
    measure3 = "(" + str(measure(collumn_source3, scale3, letter_source3)) + ")" + "*" + str(weight3/weightTotal) + "+"
    measure4 = "(" + str(measure(collumn_source4, scale4, letter_source4)) + ")" + "*" + str(weight4/weightTotal)
    count_formula = str(eq) + measure1 + measure2 + measure3 + measure4
    ws[f'{letter_target}{i+2}'].value = count_formula


# Save file 1
wb.save('Hasil SPK Mahasiswa Demotivasi.xlsx')
time.sleep(3)

#Open-close file agar file dapat dibaca oleh Openpyxl
dirOC = os.getcwd()
excelOC = win32com.client.Dispatch("Excel.Application")

wbOC = excelOC.Workbooks.Open(f'{dirOC}/Hasil SPK Mahasiswa Demotivasi.xlsx')
wsOC = wbOC.Worksheets('RAW Data')
wbOC.SaveAs(f'{dirOC}/Hasil SPK Mahasiswa Demotivasi.xlsx')
excelOC.Quit()

# Mengganti data formula ke value
wbValue = openpyxl.load_workbook('Hasil SPK Mahasiswa Demotivasi.xlsx', data_only=True)
wbValue.save('Hasil SPK Mahasiswa Demotivasi.xlsx')

print('Sistem sedang menyortir data...')
time.sleep(3)

# Sorting data
excel = win32com.client.Dispatch("Excel.Application")

wbSort = excel.Workbooks.Open(f'{dir}/Hasil SPK Mahasiswa Demotivasi.xlsx')
wsSort = wbSort.Worksheets(listSheet[0])
wsSort.Range('A2:K2000').Sort(Key1=wsSort.Range('K1'), Order1=2, Orientation=1)
wbSort.SaveAs(f'{dir}/Hasil SPK Mahasiswa Demotivasi.xlsx')
excel.Quit()

input('Penghitungan sukses! Silakan buka file "Hasil SPK Mahasiswa Demotivasi.xlsx"')


